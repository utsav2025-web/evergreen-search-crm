"""
Financial Model routes.

Endpoints
---------
GET    /financials/                     List financial rows (filter by company)
POST   /financials/                     Create a single row
GET    /financials/{id}                 Get one row
PATCH  /financials/{id}                 Update one row
DELETE /financials/{id}                 Delete one row
POST   /financials/import               Import from Excel/CSV upload
POST   /financials/ai-build             AI-build model from CIM extracts
GET    /financials/summary/{company_id} Aggregated KPI summary for a company
"""
import io
import json
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user, get_db, Pagination, require_write
from app.models.models import CIMExtract, Company, Financial, User

router = APIRouter()


def _financial_to_dict(f: Financial) -> Dict[str, Any]:
    return {
        "id": f.id,
        "company_id": f.company_id,
        "period_year": f.period_year,
        "period_type": f.period_type,
        "period_quarter": f.period_quarter,
        "revenue": f.revenue,
        "gross_profit": f.gross_profit,
        "ebitda": f.ebitda,
        "ebit": f.ebit,
        "net_income": f.net_income,
        "total_assets": f.total_assets,
        "total_debt": f.total_debt,
        "capex": f.capex,
        "owner_comp": f.owner_comp,
        "sde": f.sde,
        "source": f.source,
        "raw_data": f.raw_data or {},
        "created_at": f.created_at.isoformat() if f.created_at else None,
        "updated_at": f.updated_at.isoformat() if f.updated_at else None,
    }


@router.get("/")
async def list_financials(
    company_id: Optional[int] = Query(None),
    period_type: Optional[str] = Query(None),
    pagination: Pagination = Depends(),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
) -> Dict[str, Any]:
    q = select(Financial)
    if company_id is not None:
        q = q.where(Financial.company_id == company_id)
    if period_type:
        q = q.where(Financial.period_type == period_type)
    q = q.order_by(Financial.period_year.desc())

    total = (await db.execute(select(func.count()).select_from(q.subquery()))).scalar() or 0
    rows = (await db.execute(q.offset(pagination.skip).limit(pagination.limit))).scalars().all()
    return {"items": [_financial_to_dict(r) for r in rows], "total": total}


@router.post("/", status_code=status.HTTP_201_CREATED)
async def create_financial(
    body: Dict[str, Any],
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_write),
) -> Dict[str, Any]:
    company_id = body.get("company_id")
    if not company_id:
        raise HTTPException(400, "company_id is required")
    company = await db.get(Company, company_id)
    if not company:
        raise HTTPException(404, "Company not found")

    row = Financial(
        company_id=company_id,
        period_year=body.get("period_year", 2024),
        period_type=body.get("period_type", "annual"),
        period_quarter=body.get("period_quarter"),
        revenue=body.get("revenue"),
        gross_profit=body.get("gross_profit"),
        ebitda=body.get("ebitda"),
        ebit=body.get("ebit"),
        net_income=body.get("net_income"),
        total_assets=body.get("total_assets"),
        total_debt=body.get("total_debt"),
        capex=body.get("capex"),
        owner_comp=body.get("owner_comp"),
        sde=body.get("sde"),
        source=body.get("source", "manual"),
        raw_data=body.get("raw_data", {}),
    )
    db.add(row)
    await db.commit()
    await db.refresh(row)
    return _financial_to_dict(row)


@router.get("/summary/{company_id}")
async def financial_summary(
    company_id: int,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
) -> Dict[str, Any]:
    company = await db.get(Company, company_id)
    if not company:
        raise HTTPException(404, "Company not found")

    rows = (await db.execute(
        select(Financial)
        .where(Financial.company_id == company_id)
        .order_by(Financial.period_year.asc())
    )).scalars().all()

    if not rows:
        return {"company_id": company_id, "rows": [], "kpis": {}}

    latest = next(
        (r for r in reversed(rows) if r.period_type in ("annual", "ttm")),
        rows[-1]
    )

    annual = [r for r in rows if r.period_type == "annual" and r.revenue]
    cagr = None
    if len(annual) >= 2:
        first, last = annual[0], annual[-1]
        years = last.period_year - first.period_year
        if years > 0 and first.revenue and first.revenue > 0:
            cagr = round(((last.revenue / first.revenue) ** (1 / years) - 1) * 100, 1)

    ebitda_margin = None
    if latest.ebitda is not None and latest.revenue and latest.revenue > 0:
        ebitda_margin = round(latest.ebitda / latest.revenue * 100, 1)

    gross_margin = None
    if latest.gross_profit is not None and latest.revenue and latest.revenue > 0:
        gross_margin = round(latest.gross_profit / latest.revenue * 100, 1)

    debt_ebitda = None
    if latest.total_debt is not None and latest.ebitda and latest.ebitda > 0:
        debt_ebitda = round(latest.total_debt / latest.ebitda, 2)

    return {
        "company_id": company_id,
        "company_name": company.name,
        "rows": [_financial_to_dict(r) for r in rows],
        "kpis": {
            "latest_revenue": latest.revenue,
            "latest_ebitda": latest.ebitda,
            "latest_sde": latest.sde,
            "ebitda_margin_pct": ebitda_margin,
            "gross_margin_pct": gross_margin,
            "revenue_cagr_pct": cagr,
            "debt_ebitda": debt_ebitda,
            "latest_period": f"{latest.period_year} {latest.period_type}",
        },
    }


@router.get("/{financial_id}")
async def get_financial(
    financial_id: int,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
) -> Dict[str, Any]:
    row = await db.get(Financial, financial_id)
    if not row:
        raise HTTPException(404, "Financial record not found")
    return _financial_to_dict(row)


@router.patch("/{financial_id}")
async def update_financial(
    financial_id: int,
    body: Dict[str, Any],
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_write),
) -> Dict[str, Any]:
    row = await db.get(Financial, financial_id)
    if not row:
        raise HTTPException(404, "Financial record not found")

    updatable = [
        "period_year", "period_type", "period_quarter",
        "revenue", "gross_profit", "ebitda", "ebit", "net_income",
        "total_assets", "total_debt", "capex", "owner_comp", "sde",
        "source", "raw_data",
    ]
    for field in updatable:
        if field in body:
            setattr(row, field, body[field])

    await db.commit()
    await db.refresh(row)
    return _financial_to_dict(row)


@router.delete("/{financial_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_financial(
    financial_id: int,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_write),
):
    row = await db.get(Financial, financial_id)
    if not row:
        raise HTTPException(404, "Financial record not found")
    await db.delete(row)
    await db.commit()


@router.post("/import")
async def import_financials(
    company_id: int = Query(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_write),
) -> Dict[str, Any]:
    """Import from Excel (.xlsx) or CSV file."""
    company = await db.get(Company, company_id)
    if not company:
        raise HTTPException(404, "Company not found")

    content = await file.read()
    filename = file.filename or ""

    try:
        if filename.endswith(".csv"):
            import csv
            reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
            data_rows = list(reader)
        elif filename.endswith((".xlsx", ".xls")):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
            data_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                data_rows.append(dict(zip(headers, row)))
        else:
            raise HTTPException(422, "Only .csv, .xlsx, or .xls files are supported")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(422, f"Could not parse file: {e}")

    numeric_fields = [
        "revenue", "gross_profit", "ebitda", "ebit", "net_income",
        "total_assets", "total_debt", "capex", "owner_comp", "sde",
    ]

    created = 0
    updated = 0
    errors: List[str] = []

    for i, row in enumerate(data_rows):
        try:
            year = int(row.get("period_year") or row.get("year") or 0)
            if not year:
                errors.append(f"Row {i+2}: missing period_year")
                continue
            ptype = str(row.get("period_type") or "annual").lower()
            pquarter = row.get("period_quarter") or row.get("quarter")
            pquarter = int(pquarter) if pquarter else None

            existing = (await db.execute(
                select(Financial).where(
                    Financial.company_id == company_id,
                    Financial.period_year == year,
                    Financial.period_type == ptype,
                    Financial.period_quarter == pquarter,
                )
            )).scalar_one_or_none()

            vals: Dict[str, Any] = {}
            for f in numeric_fields:
                v = row.get(f)
                if v is not None and str(v).strip() not in ("", "None", "N/A", "-"):
                    try:
                        vals[f] = float(str(v).replace(",", "").replace("$", ""))
                    except ValueError:
                        pass

            if existing:
                for k, v in vals.items():
                    setattr(existing, k, v)
                existing.source = "excel_import"
                updated += 1
            else:
                new_row = Financial(
                    company_id=company_id,
                    period_year=year,
                    period_type=ptype,
                    period_quarter=pquarter,
                    source="excel_import",
                    raw_data=dict(row),
                    **vals,
                )
                db.add(new_row)
                created += 1
        except Exception as e:
            errors.append(f"Row {i+2}: {e}")

    await db.commit()
    return {
        "status": "ok",
        "created": created,
        "updated": updated,
        "errors": errors,
        "total_rows": len(data_rows),
    }


@router.post("/ai-build")
async def ai_build_model(
    company_id: int = Query(...),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_write),
) -> Dict[str, Any]:
    """Use AI to extract structured financial rows from CIM extract."""
    company = await db.get(Company, company_id)
    if not company:
        raise HTTPException(404, "Company not found")

    extract = (await db.execute(
        select(CIMExtract)
        .where(CIMExtract.company_id == company_id)
        .order_by(CIMExtract.created_at.desc())
        .limit(1)
    )).scalar_one_or_none()

    if not extract:
        raise HTTPException(404, "No CIM extract found. Upload and parse a CIM first.")

    summary = extract.financials_summary or {}
    raw_text_snippet = (extract.raw_text or "")[:3000]

    prompt = f"""You are a financial analyst. Extract structured annual financial data from the following CIM information.

Company: {company.name}
Industry: {company.industry or 'Unknown'}

Financial Summary from CIM:
{json.dumps(summary, indent=2)}

Raw text excerpt:
{raw_text_snippet}

Return a JSON array of financial period objects. Each object must have these fields (use null if unknown):
- period_year (integer, e.g. 2023)
- period_type (string: "annual", "ttm", or "quarterly")
- period_quarter (integer 1-4 or null)
- revenue (float in dollars, e.g. 5000000)
- gross_profit (float or null)
- ebitda (float or null)
- ebit (float or null)
- net_income (float or null)
- owner_comp (float or null)
- sde (float or null)
- total_debt (float or null)
- capex (float or null)

Include at minimum the TTM/most recent year. Include up to 5 years if available.
Return ONLY valid JSON array, no explanation."""

    try:
        from openai import OpenAI
        import os
        client = OpenAI(
            api_key=os.environ.get("OPENAI_API_KEY"),
            base_url=os.environ.get("OPENAI_BASE_URL", "https://api.openai.com/v1"),
        )
        resp = client.chat.completions.create(
            model="gpt-4.1-mini",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.1,
            max_tokens=2000,
        )
        raw = resp.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        financial_rows = json.loads(raw)
    except Exception as e:
        raise HTTPException(500, f"AI extraction failed: {e}")

    created = 0
    updated = 0
    for row_data in financial_rows:
        try:
            year = int(row_data.get("period_year") or 0)
            if not year:
                continue
            ptype = str(row_data.get("period_type") or "annual").lower()
            pquarter = row_data.get("period_quarter")

            existing = (await db.execute(
                select(Financial).where(
                    Financial.company_id == company_id,
                    Financial.period_year == year,
                    Financial.period_type == ptype,
                    Financial.period_quarter == pquarter,
                )
            )).scalar_one_or_none()

            numeric_fields = [
                "revenue", "gross_profit", "ebitda", "ebit", "net_income",
                "total_assets", "total_debt", "capex", "owner_comp", "sde",
            ]
            vals = {f: row_data.get(f) for f in numeric_fields if row_data.get(f) is not None}

            if existing:
                for k, v in vals.items():
                    if getattr(existing, k) is None:
                        setattr(existing, k, v)
                existing.source = "cim_extract"
                updated += 1
            else:
                new_row = Financial(
                    company_id=company_id,
                    period_year=year,
                    period_type=ptype,
                    period_quarter=pquarter,
                    source="cim_extract",
                    raw_data=row_data,
                    **vals,
                )
                db.add(new_row)
                created += 1
        except Exception:
            continue

    await db.commit()
    return {
        "status": "ok",
        "created": created,
        "updated": updated,
        "ai_rows": financial_rows,
    }
